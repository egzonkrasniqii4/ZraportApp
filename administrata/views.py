#administrata/ views.py
import base64
from datetime import datetime
from django.shortcuts import render,redirect
import pyodbc
from datetime import datetime, timedelta
import urllib.parse
import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from django.http import HttpResponse
import pyodbc
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse


conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=192.168.49.49;'
    'DATABASE=ZraportsApp;'
    'UID=sa;'
    'PWD=sasa'
)
# Main Page View
def index(request):
    return render(request, 'administrata/index.html')

# Albi Fashion Reports
def albi_fashion(request):
    return render(request, 'administrata/albi_fashion.html')


@csrf_exempt  # If you need to disable CSRF protection for this view
def procedura_albi_fashion(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[01]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[01] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '01'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[01]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[01]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[01] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '01'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('011', 'KON-0001'), ('012', 'KON-0002'), ('017', 'KON-0003'), ('034', 'KON-0004'), 
                                    ('035', 'KON-0005'), ('039', 'KON-0006'), ('040', 'KON-0007'), ('045', 'KON-0008'),
                                    ('046', 'KON-0009'), ('048', 'KON-0011'), ('049', 'KON-0012'), ('051', 'KON-0014'),
                                    ('052', 'KON-0015'), ('053', 'KON-0016'), ('054', 'KON-0017'), ('055', 'KON-0019'), 
                                    ('065', 'KON-0030'), ('060', 'KON-0033'), ('061', 'KON-0034'), ('056', 'KON-0020'), 
                                    ('063', 'KON-0036'), ('064', 'KON-0037'), ('066', 'KON-0038'), ('068', 'KON-0039'), 
                                    ('067', 'KON-0040'), ('069', 'KON-0041'), ('070', 'KON-0042'), ('072', 'KON-0044'),
                                    ('073', 'KON-0045'), ('074', 'KON-0046')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[01]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[01] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '01';
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                #  JSON response success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                #  JSON response error 
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/01/procedura_albi_fashion.html')





def edit_z_report_albi_fashion(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [01]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_albi_fashion')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Fetch the old date value before updating
                fetch_query = """
                SELECT Date 
                FROM [01]
                WHERE ImageData = ? AND Location = ?
                """
                cursor.execute(fetch_query, (image_data, location))
                record = cursor.fetchone()
                old_date = record.Date if record else None

                if old_date:
                    update_query = """
                    UPDATE [01]
                    SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                    WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                    """

                    # Convert None values to SQL NULL
                    sistemi = None if sistemi in [None, 'None', ''] else sistemi
                    difference = None if difference in [None, 'None', ''] else difference

                    cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                    conn.commit()

                    return redirect('daily_z_report_albi_fashion')
                else:
                    error_message = 'Record not found to update.'

            except Exception as e:
                error_message = f"Error updating record: {e}"

            finally:
                cursor.close()
                conn.close()

    else:  # Handle GET request
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [01]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/01/edit_z_report_albi_fashion.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/01/edit_z_report_albi_fashion.html', {
        'error_message': error_message
    })


def daily_z_report_albi_fashion(request):
    # SQL Server connection string

    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-01];")  
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [01] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/01/daily_z_report_albi_fashion.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })



def monthly_z_report_albi_fashion(request):

   
    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/01/monthly_z_report_albi_fashion.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_01];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '01'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/01/monthly_z_report_albi_fashion.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )




def ardi_co(request):
    return render(request, 'administrata/ardi_co.html')

def procedura_ardi_co(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[02]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[02] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '02'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[02]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[02]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[02] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '02'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('043', '043')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[02]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[02] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '02';
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/02/procedura_ardi_co.html')





def edit_z_report_ardi_co(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [02]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_ardi_co')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Fetch the old date value before updating
            fetch_query = """
            SELECT Date 
            FROM [02]
            WHERE ImageData = ? AND Location = ?
            """
            cursor.execute(fetch_query, (image_data, location))
            record = cursor.fetchone()
            old_date = record.Date if record else None

            if old_date:
                update_query = """
                UPDATE [02]
                SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """

                sistemi = None if sistemi in [None, 'None', ''] else sistemi
                difference = None if difference in [None, 'None', ''] else difference

                cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                conn.commit()

                return redirect('daily_z_report_ardi_co')
            else:
                error_message = 'Record not found to update.'

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [02]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/02/edit_z_report_ardi_co.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/02/edit_z_report_ardi_co.html', {
        'error_message': error_message
    })



def daily_z_report_ardi_co(request):
    # SQL Server connection string

    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-02];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,  
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [02] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/02/daily_z_report_ardi_co.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })

def monthly_z_report_ardi_co(request):

    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/01/monthly_z_report_ardi_co.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_02];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '02'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/02/monthly_z_report_ardi_co.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )

def nan(request):
    return render(request, 'administrata/nan.html')

def procedura_nan(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[03]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[03] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '03'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[03]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[03]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[03] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '03'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('0311', '0311'), ('0312', '0312'), ('0313', '0313'), ('0314', '0314'), ('0315','0315')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[03]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[03] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '03';
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/03/procedura_nan.html')





def edit_z_report_nan(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [03]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_nan')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Fetch the old date value before updating
                fetch_query = """
                SELECT Date 
                FROM [03]
                WHERE ImageData = ? AND Location = ?
                """
                cursor.execute(fetch_query, (image_data, location))
                record = cursor.fetchone()
                old_date = record.Date if record else None

                if old_date:
                    update_query = """
                    UPDATE [03]
                    SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                    WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                    """

                    # Convert None values to SQL NULL
                    sistemi = None if sistemi in [None, 'None', ''] else sistemi
                    difference = None if difference in [None, 'None', ''] else difference

                    cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                    conn.commit()

                    return redirect('daily_z_report_nan')
                else:
                    error_message = 'Record not found to update.'

            except Exception as e:
                error_message = f"Error updating record: {e}"

            finally:
                cursor.close()
                conn.close()

    else:  # Handle GET request
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card,
                   Cupon,
                   Other,   
                   Difference,
                   ImageData
            FROM [03]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/03/edit_z_report_nan.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/03/edit_z_report_nan.html', {
        'error_message': error_message
    })



def daily_z_report_nan(request):


    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-03];")  
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,  
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [03] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/03/daily_z_report_nan.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })




def monthly_z_report_nan(request):

    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/03/monthly_z_report_nan.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_03];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '03'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/03/monthly_z_report_nan.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )

def obe(request):
    return render(request, 'administrata/obe.html')


def procedura_obe(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[04]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[04] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '04'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[04]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[04]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[04] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '04'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('0413', '0413'), ('0414', '0414')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[04]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[04] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '04';
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/04/procedura_obe.html')






def edit_z_report_obe(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [04]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_obe')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Fetch the old date value before updating
            fetch_query = """
            SELECT Date 
            FROM [04]
            WHERE ImageData = ? AND Location = ?
            """
            cursor.execute(fetch_query, (image_data, location))
            record = cursor.fetchone()
            old_date = record.Date if record else None

            if old_date:
                update_query = """
                UPDATE [04]
                SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """

                sistemi = None if sistemi in [None, 'None', ''] else sistemi
                difference = None if difference in [None, 'None', ''] else difference


                cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                conn.commit()

                return redirect('daily_z_report_obe')
            else:
                error_message = 'Record not found to update.'

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [04]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/04/edit_z_report_obe.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/04/edit_z_report_obe.html', {
        'error_message': error_message
    })



def daily_z_report_obe(request):


    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-04];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [04] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/04/daily_z_report_obe.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })






def monthly_z_report_obe(request):
    

    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/04/monthly_z_report_obe.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_04];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '04'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/04/monthly_z_report_obe.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )

def dndo(request):
    return render(request, 'administrata/dndo.html')


def procedura_dndo(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[05]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[05] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '05'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[05]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[05]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[05] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '05'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('0511', '0511'), ('0514', '0514'), ('0515', '0515'), ('0516', '0516'), 
                                    ('0517', '0517'), ('0520', '0520'), ('0521', '0521'), ('0523', '0523'),('0524', '0524')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[05]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[05] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '05';

                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/05/procedura_dndo.html')

def edit_z_report_dndo(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [05]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_dndo')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Fetch the old date value before updating
            fetch_query = """
            SELECT Date 
            FROM [05]
            WHERE ImageData = ? AND Location = ?
            """
            cursor.execute(fetch_query, (image_data, location))
            record = cursor.fetchone()
            old_date = record.Date if record else None

            if old_date:
                update_query = """
                UPDATE [05]
                SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """


                sistemi = None if sistemi in [None, 'None', ''] else sistemi
                difference = None if difference in [None, 'None', ''] else difference


                cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                conn.commit()

                return redirect('daily_z_report_dndo')
            else:
                error_message = 'Record not found to update.'

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [05]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/05/edit_z_report_dndo.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/05/edit_z_report_dndo.html', {
        'error_message': error_message
    })



def daily_z_report_dndo(request):


    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-05];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [05] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/05/daily_z_report_dndo.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })






def monthly_z_report_dndo(request):
    
    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/05/monthly_z_report_dndo.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_05];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '05'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/05/monthly_z_report_dndo.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )    
def jaroma(request):
    return render(request, 'administrata/jaroma.html')


def procedura_jaroma(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[06]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[06] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '06'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[06]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[06]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[06] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '06'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('0611', '0611'), ('0612','0612')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[06]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[06] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '06';
                
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/06/procedura_jaroma.html')


def edit_z_report_jaroma(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [06]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_jaroma')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')

        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Fetch the old date value before updating
            fetch_query = """
            SELECT Date 
            FROM [06]
            WHERE ImageData = ? AND Location = ?
            """
            cursor.execute(fetch_query, (image_data, location))
            record = cursor.fetchone()
            old_date = record.Date if record else None

            if old_date:
                update_query = """
                UPDATE [06]
                SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                
                sistemi = None if sistemi in [None, 'None', ''] else sistemi
                difference = None if difference in [None, 'None', ''] else difference


                cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                conn.commit()

                return redirect('daily_z_report_jaroma')
            else:
                error_message = 'Record not found to update.'

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [06]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/06/edit_z_report_jaroma.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/06/edit_z_report_jaroma.html', {
        'error_message': error_message
    })



def daily_z_report_jaroma(request):


    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-06];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [06] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/06/daily_z_report_jaroma.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })


def monthly_z_report_jaroma(request):
    

    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/06/monthly_z_report_jaroma.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_06];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '06'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/06/monthly_z_report_jaroma.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )

def albi_fashion_retail(request):
    return render(request, 'administrata/albi_fashion_retail.html')

def procedura_albi_fashion_retail(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[07]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[07] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '07'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[07]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[07]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[07] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '07'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('2570', '2570'), ('2580', '2580'), ('2615', '2615'), ('2620', '2620'),
                                    ('2626', '2626'), ('2627', '2627'), ('2631', '2631'), ('2633', '2633'),
                                    ('2636', '2636'), ('2637', '2637'), ('2638', '2638'), ('2641', '2641'),
                                    ('2642', '2642'), ('2643', '2643'), ('2660', '2660'), ('2662', '2662'), 
                                    ('2663', '2663'), ('2675', '2675')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[07]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[07] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '07';
                
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/07/procedura_albi_fashion_retail.html')



def edit_z_report_albi_fashion_retail(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [07]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_albi_fashion_retail')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Fetch the old date value before updating
            fetch_query = """
            SELECT Date 
            FROM [07]
            WHERE ImageData = ? AND Location = ?
            """
            cursor.execute(fetch_query, (image_data, location))
            record = cursor.fetchone()
            old_date = record.Date if record else None

            if old_date:
                update_query = """
                UPDATE [07]
                SET Date = ?, TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """

                sistemi = None if sistemi in [None, 'None', ''] else sistemi
                difference = None if difference in [None, 'None', ''] else difference


                cursor.execute(update_query, (new_date, total_value, sistemi, cash, card, cupon, difference, image_data, location, old_date))
                conn.commit()

                return redirect('daily_z_report_albi_fashion_retail')
            else:
                error_message = 'Record not found to update.'

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [07]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/07/edit_z_report_albi_fashion_retail.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/07/edit_z_report_albi_fashion_retail.html', {
        'error_message': error_message
    })



def daily_z_report_albi_fashion_retail(request):
    # SQL Server connection string

    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-07];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [07] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/07/daily_z_report_albi_fashion_retail.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })





def monthly_z_report_albi_fashion_retail(request):
    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/07/monthly_z_report_albi_fashion_retail.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_07];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '07'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/07/monthly_z_report_albi_fashion_retail.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )

def ran(request):
    return render(request, 'administrata/ran.html')


def procedura_ran(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # Execute your SQL queries with the selected date
                query = f"""
                DECLARE @SelectedDate DATE = '{selected_date}';

                -- Step 1: Update the Sistemi column
                UPDATE [ZRaportsApp].[dbo].[08]
                SET [Sistemi] = ISNULL(SalesData.Sistemi, 0)
                FROM [ZRaportsApp].[dbo].[08] AS Target
                INNER JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Sistemi,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '08'
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS SalesData
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = SalesData.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = SalesData.[Posting Date];

                -- Step 2: Update the Difference column
                UPDATE [ZRaportsApp].[dbo].[08]
                SET [Difference] = [TotalValue] - [Sistemi] 
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 3: Update the Other column
                UPDATE [ZRaportsApp].[dbo].[08]
                SET [Other] = ISNULL(UnlinkedSales.Other, 0)
                FROM [ZRaportsApp].[dbo].[08] AS Target
                LEFT JOIN (
                    SELECT 
                        [Location Code] AS LocationCode,
                        SUM([Amount Including VAT]) AS Other,
                        [Posting Date]
                    FROM 
                        [Fashion POS].[dbo].[BI_FactSalesDetailsParagons]
                    WHERE 
                        [Posting Date] = @SelectedDate AND [Company Code] = '08'
                        AND [Sell-to Customer No_] != '8888'
                        AND NOT EXISTS (
                            SELECT 1
                            FROM (
                                VALUES 
                                    ('011', '011'), ('012', '012')
                            ) AS Link (LocationCode, CustomerNo)
                            WHERE Link.LocationCode = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Location Code]
                            AND Link.CustomerNo = [Fashion POS].[dbo].[BI_FactSalesDetailsParagons].[Sell-to Customer No_]
                        )
                    GROUP BY 
                        [Location Code], [Posting Date]
                ) AS UnlinkedSales
                ON Target.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = UnlinkedSales.LocationCode COLLATE SQL_Latin1_General_CP1_CI_AS
                AND CONVERT(DATE, Target.[Date]) = UnlinkedSales.[Posting Date]
                WHERE CONVERT(DATE, [Date]) = @SelectedDate;

                -- Step 4: Update the Store column
                UPDATE [ZRaportsApp].[dbo].[08]
                SET [Store] = B.[Location Name]
                FROM [ZRaportsApp].[dbo].[08] AS A
                INNER JOIN [Fashion POS].[dbo].[BI_DimLocation] AS B
                ON A.[Location] COLLATE SQL_Latin1_General_CP1_CI_AS = B.[Location Code] COLLATE SQL_Latin1_General_CP1_CI_AS
                WHERE B.[Company Code] = '08';
                
                """
                cursor.execute(query)
                conn.commit()
                conn.close()
                # Return JSON response indicating success
                return JsonResponse({"message": "Procedura u ekzekutua me sukses"}, status=200)
            except Exception as e:
                # Return JSON response with error message
                return JsonResponse({"error": str(e)}, status=500)

    return render(request, 'administrata/08/procedura_ran.html')

def edit_z_report_ran(request):
    error_message = None  # Initialize error message variable

    if request.method == 'POST':
        if 'delete' in request.POST:  # Check if the delete button was clicked
            location = request.POST.get('location')
            image_data = request.POST.get('image_data')
            date = request.POST.get('date')

            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                delete_query = """
                DELETE FROM [08]
                WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
                """
                cursor.execute(delete_query, (image_data, location, date))
                conn.commit()

                return redirect('daily_z_report_ran')

            except Exception as e:
                error_message = f"Error deleting record: {e}"

            finally:
                cursor.close()
                conn.close()

        else:  # Handle the update operation
            location = request.POST.get('location')
            new_date = request.POST.get('date')  # Get the new date value from the form
            image_data = request.POST.get('image_data')
            total_value = request.POST.get('total_value')
            sistemi = request.POST.get('sistemi')
            cash = request.POST.get('cash')
            card = request.POST.get('card')
            cupon = request.POST.get('cupon')
            difference = request.POST.get('difference')

        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            update_query = """
            UPDATE [08]
            SET TotalValue = ?, Sistemi = ?, Cash = ?, Card = ?, Cupon = ?, Difference = ?, EditFlag = 1
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """

            sistemi = None if sistemi in [None, 'None', ''] else sistemi
            difference = None if difference in [None, 'None', ''] else difference

            
            cursor.execute(update_query, (total_value, sistemi, cash, card, cupon, difference, image_data, location, date))
            conn.commit()

            return redirect('daily_z_report_jaroma')

        except Exception as e:
            error_message = f"Error updating record: {e}"

        finally:
            cursor.close()
            conn.close()

    else:
        image_data = request.GET.get('image_data')
        location = request.GET.get('location')
        date = request.GET.get('date')

        # Decode image_data if necessary
        if image_data:
            image_data = urllib.parse.unquote(image_data)
            # Extract only the relative path part of image_data if it contains a URL
            if 'http://' in image_data or 'https://' in image_data:
                image_data = image_data.split('/')[-1]

        # Debugging output
        print(f"Decoded ImageData: {image_data}, Location: {location}, Date: {date}")

        if image_data:
            query = """
            SELECT [ID],
                   Location, 
                   CONVERT(varchar, Date, 23) as Date, 
                   TotalValue,
                   Sistemi,
                   Cash, 
                   Card, 
                   Cupon, 
                   Other,        
                   Difference,
                   ImageData
            FROM [08]
            WHERE ImageData = ? AND Location = ? AND CONVERT(varchar, Date, 23) = ?
            """
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute(query, (image_data, location, date))
                record = cursor.fetchone()

                # Debugging output
                print(f"Fetched record: {record}")

                if record:
                    return render(request, 'administrata/08/edit_z_report_jaroma.html', {
                        'record': record,
                        'error_message': error_message
                    })
                else:
                    error_message = 'Record not found. Please try again.'

            except Exception as e:
                error_message = f"Error retrieving record: {e}"

            finally:
                cursor.close()
                conn.close()
        else:
            error_message = 'Invalid parameters provided.'

    return render(request, 'administrata/08/edit_z_report_ran.html', {
        'error_message': error_message
    })



def daily_z_report_ran(request):


    # Default date is yesterday
    today = datetime.now()
    default_date = (today - timedelta(days=1)).date()  # Default is yesterday
    selected_start_date = request.GET.get('start_date', default_date)
    selected_end_date = request.GET.get('end_date', default_date)
    selected_location = request.GET.get('location', '')
    selected_difference = request.GET.get('difference', '')


    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForYesterday-08]")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()

    # Prepare SQL query with placeholders
    query = """
    SELECT Store,
           Location,   
           CONVERT(varchar, Date, 23) as Date, 
           TotalValue,
           Sistemi,
           Cash, 
           Card, 
           Cupon, 
           Other,
           Difference,
           ImageData,
           EditFlag
    FROM [08] 
    WHERE 1=1
    """
    params = []

    # Add filters based on user input or default behavior
    if selected_start_date and selected_end_date:
        query += " AND Date BETWEEN ? AND ?"
        params.append(selected_start_date)
        params.append(selected_end_date)

    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"  # Include all locations if not specified

    if selected_difference == '0':  # "No" means Difference = 0
        query += " AND CAST([Difference] AS INT) = 0"
    elif selected_difference == '1':  # "Yes" means Difference <> 0
        query += " AND (CAST([Difference] AS INT) <> 0 OR [Difference] IS NULL)"

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))  # Execute the query with parameters

        # Fetch all records
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]  # Get column names
        for row in records:
            record = dict(zip(columns, row))  # Create a dictionary for the row
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

    except Exception as e:
        print(f"Error: {e}")

    finally:
        cursor.close()
        conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Daily_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    return render(request, 'administrata/08/daily_z_report_ran.html', {
        'records': records_list,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_location': selected_location,
        'selected_difference': selected_difference,
        'today': today.date(),  # Pass today's date
    })



def monthly_z_report_ran(request):
    # Default values for the previous month
    today = datetime.now()
    previous_month = today.replace(day=1) - timedelta(days=1)  # Move to the last day of the previous month
    default_month = previous_month.strftime('%Y-%m')
    
    # Retrieve form inputs
    selected_from_month = request.GET.get('from_month', default_month)
    selected_to_month = request.GET.get('to_month', default_month)
    selected_location = request.GET.get('location', '').strip()

    # Log selected months and location
    print(f"From Month: {selected_from_month}, To Month: {selected_to_month}, Location: {selected_location}")

    # Extract year and month separately
    try:
        from_year, from_month = map(int, selected_from_month.split('-'))
        to_year, to_month = map(int, selected_to_month.split('-'))
    except ValueError:
        print("Invalid date format")
        return render(
            request,
            'administrata/08/monthly_z_report_ran.html',
            {
                'records': [],
                'selected_from_month': selected_from_month,
                'selected_to_month': selected_to_month,
                'selected_location': selected_location,
                'today': today,
                'error': "Invalid date format."
            }
        )

    # Construct date range as integer
    from_period = from_year * 100 + from_month
    to_period = to_year * 100 + to_month

    # Base SQL query

    if request.GET.get('run_procedure') == 'true':
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("EXEC [UpdateAllSalesTotalsForPreviousMonth_08];")  # Execute stored procedure
            conn.commit()  # Commit if necessary
        except Exception as e:
            print(f"Error executing stored procedure: {e}")
        finally:
            cursor.close()
            conn.close()
    
    query = """
        SELECT Location,
               Store,
               ReceivedMonth,
               TotalValue,
               Sistemi,
               Other,
               Difference,
               ImageData
        FROM [ZPeriod]
        WHERE [COMPANY] = '08'
          AND (202500 + ReceivedMonth) BETWEEN ? AND ?
    """
    
    # Construct the parameters
    params = [from_period, to_period]

    # Add location filter if provided
    if selected_location:
        query += " AND Location LIKE ?"
        params.append(f"%{selected_location}%")
    else:
        query += " AND Location LIKE '%'"

    # Debugging output to see the query and parameters
    print(f"Final Query: {query}")
    print(f"Parameters: {params}")

    records_list = []  # Initialize the records list
    columns = []  # Initialize columns list

    try:
        # Establish database connection
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Execute query with parameters
        cursor.execute(query, params)
        records = cursor.fetchall()

        # Convert records to a list of dictionaries
        columns = [column[0] for column in cursor.description]
        for row in records:
            record = dict(zip(columns, row))
            # Convert image data to a URL if it exists
            if record['ImageData']:
                record['ImageData'] = f"http://192.168.49.3:8000/static/SCANAPP/{record['ImageData']}"
            records_list.append(record)

        if not records_list:
            print("No records found for the selected criteria.")
    except pyodbc.Error as e:
        print("Error executing query:", e)
    finally:
        if 'conn' in locals():
            conn.close()

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Monthly_Z_Report_{today.date()}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Z Report"

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(records_list, 2):
            for col_num, column_title in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = record[column_title]

        wb.save(response)
        return response

    # Render the response
    return render(
        request,
        'administrata/08/monthly_z_report_ran.html',
        {
            'records': records_list,
            'selected_from_month': selected_from_month,
            'selected_to_month': selected_to_month,
            'selected_location': selected_location,
            'today': today
        }
    )
